import openpyxl


workbook = openpyxl.load_workbook('task.xlsx')
sheet = workbook.active



data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)


for i in range(1, len(data)):
    weight = data[i][1]
    price = data[i][2]
    year = data[i][0].year
    month = data[i][0].month
    day = data[i][0].day
    number = data[i][4]
    file_name = f"completed/Торг-12/{month}{day}-{number}  от {day}.{month}.{year}.xlsx"
    workbook2 = openpyxl.load_workbook('templates/Торг-12/0101-1  от 01.01.2023.xlsx')
    sheet2 = workbook2.active
    sheet2['w23'] = weight
    sheet2['z23'] = price
    workbook2.save(file_name)
    
